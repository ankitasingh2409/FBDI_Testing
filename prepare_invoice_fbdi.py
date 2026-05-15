# =============================================================================
# Oracle Fusion AR — Invoice FBDI Preparation Script
#
# Usage:
#   python prepare_invoice_fbdi.py                  <- process all batches found
#   python prepare_invoice_fbdi.py --group DEFAKTO_INV_0001
#   python prepare_invoice_fbdi.py --list-groups
#   python prepare_invoice_fbdi.py --parallel
#
# Every complete TRX + TRX_DIST pair found in source_data/ is processed
# automatically. One output .xlsm is produced per pair.
#
# Folder structure expected:
#   Invoice_FBDI/
#   ├── prepare_invoice_fbdi.py         <- this script
#   ├── source_config.json              <- column mappings and configuration
#   ├── source_data/
#   │   ├── TRX_DEFAKTO_INV_0001_260423_091853.csv
#   │   ├── TRX_DIST_DEFAKTO_INV_0001_260423_091853.csv
#   │   ├── TRX_GENEVA_INV_0001_...csv
#   │   └── TRX_DIST_GENEVA_INV_0001_...csv
#   ├── templates/
#   │   └── Invoice_Template.xlsm
#   └── output/
#       ├── Invoice_Template_DEFAKTO_INV_0001.xlsm
#       └── Invoice_Template_GENEVA_INV_0001.xlsm
#
# Source file naming convention:
#   TRX_{SOURCE}_{TYPE}_{SEQ}_{DATE}_{TIME}.csv
#   TRX_DIST_{SOURCE}_{TYPE}_{SEQ}_{DATE}_{TIME}.csv
#   where:
#     SOURCE = source system name  (e.g. DEFAKTO, GENEVA)
#     TYPE   = INV | CM | DM
#     SEQ    = 4-digit sequence    (e.g. 0001, 0002)
#     DATE   = DDMMYY
#     TIME   = HHMMSS
#
# Batch identifier = {SOURCE}_{TYPE}_{SEQ}  e.g. DEFAKTO_INV_0001
#
# File mapping:
#   TRX_*      -> RA_INTERFACE_LINES_ALL sheet
#   TRX_DIST_* -> RA_INTERFACE_DISTRIBUTIONS_ALL sheet
#
# Dependencies:
#   pip install pandas openpyxl pywin32   # pywin32 recommended for macro preservation
# =============================================================================

import argparse
import concurrent.futures
import json
import os
import re
import shutil
import sys
from datetime import datetime

import openpyxl
import pandas as pd


# =============================================================================
# CONFIGURATION
# =============================================================================

SOURCE_DIR        = "source_data"
TEMPLATE_DIR      = "templates"
OUTPUT_DIR        = "output"
CONFIG_FILE       = "source_config.json"
TEMPLATE_FILENAME = "Invoice_Template.xlsm"

SOURCE_KEYS = ["trx", "trx_dist"]

DEFAULT_SOURCE_FILE_PREFIXES = {
    "trx":      "TRX",
    "trx_dist": "TRX_DIST",
}

# Valid transaction type codes embedded in source filenames
VALID_TRANSACTION_TYPES = {"INV", "CM", "DM"}

# Oracle FBDI template sheet names for invoice data
SHEET_NAMES = {
    "trx":      "RA_INTERFACE_LINES_ALL",
    "trx_dist": "RA_INTERFACE_DISTRIBUTIONS_ALL",
}

# Internal linking column in source files (not mapped to the template)
TRX_ID_COLUMN = "trx_id"


def _controlled_trx_template_columns():
    """TRX template columns treated as source-specific flexfield families."""
    def _norm(name):
        return "".join(ch.lower() for ch in str(name).strip() if ch.isalnum())

    controlled = {"Line Transactions Flexfield Context", "Link-to Transactions Flexfield Context"}
    for i in range(1, 16):
        controlled.add(f"Line Transactions Flexfield Segment {i}")
        controlled.add(f"Link-to Transactions Flexfield Segment {i}")
    return {_norm(col) for col in controlled}


def _controlled_trx_dist_template_columns():
    """TRX_DIST template columns treated as source-specific flexfield families."""
    def _norm(name):
        return "".join(ch.lower() for ch in str(name).strip() if ch.isalnum())

    controlled = {"Line Transactions Flexfield Context"}
    for i in range(1, 16):
        controlled.add(f"Line Transactions Flexfield Segment {i}")
    return {_norm(col) for col in controlled}


CONTROLLED_TRX_TEMPLATE_COLUMNS = _controlled_trx_template_columns()
CONTROLLED_TRX_DIST_TEMPLATE_COLUMNS = _controlled_trx_dist_template_columns()


# =============================================================================
# HELPERS
# =============================================================================

def _excel_col_letter(n):
    """Convert 1-based column number to Excel letter (1 -> A, 27 -> AA)."""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def format_context(group_id=None, file_path=None, sheet_name=None, key=None):
    parts = []
    if group_id:
        parts.append(f"group={group_id}")
    if file_path:
        parts.append(f"file={os.path.basename(file_path)}")
    if sheet_name:
        parts.append(f"sheet={sheet_name}")
    elif key:
        parts.append(f"sheet_key={key}")
    return " ".join(parts) if parts else None


def log(msg, level="INFO", context=None):
    prefix = {
        "INFO":   "  ",
        "OK":     "  [OK]",
        "WARN":   "  [WARN]",
        "ERROR":  "  [ERROR]",
        "HEADER": "",
    }.get(level, "  ")
    ts = datetime.now().strftime("%H:%M:%S")
    ctx = f" [{context}]" if context else ""
    print(f"[{ts}] {prefix}{ctx} {msg}")


def clean_str(val):
    """Return normalized string; empty string for NaN/None only."""
    if isinstance(val, float) and pd.isna(val):
        return ""
    if val is None:
        return ""
    return " ".join(str(val).strip().split())


def normalize_column_name(name):
    """Normalize a column header: lowercase, alphanumeric characters only."""
    if name is None:
        return ""
    return "".join(ch.lower() for ch in str(name).strip() if ch.isalnum())


def is_date_column(header_name):
    """Return True if the template column name indicates a date field."""
    if not header_name:
        return False
    return bool(re.search(r"\bDate\b", str(header_name), flags=re.IGNORECASE))


def is_text_code_column(header_name):
    """Return True for template columns that should always be written as text."""
    if not header_name:
        return False
    normalized = normalize_column_name(header_name)
    return bool(
        normalized.startswith("linetransactionsflexfieldsegment")
        or normalized == "linetransactionsflexfieldcontext"
        or normalized.startswith("linktotransactionsflexfieldsegment")
        or normalized == "linktotransactionsflexfieldcontext"
        or normalized.startswith("accountingflexfieldsegment")
    )


def _has_leading_zero_values(series):
    """Return True if any non-empty value is a digit-only string with a leading zero.

    Excel silently collapses such strings to 0 unless the cell is text-formatted.
    """
    for v in series:
        s = clean_str(v)
        if len(s) > 1 and s[0] == "0" and s.isdigit():
            return True
    return False


_DATE_FORMAT_RE = re.compile(r"^\d{4}/\d{2}/\d{2}$")


def parse_to_date_value(value):
    """
    Parse a date string and return it formatted as YYYY/MM/DD.
    Handles Oracle formats such as '05-JUN-25' and 'DD-MON-YYYY'.
    Returns the formatted string, the original string if unparseable, or None if empty.
    """
    if isinstance(value, float) and pd.isna(value):
        return None
    if value is None:
        return None
    value = clean_str(str(value))
    if not value:
        return None
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return value
    return parsed.strftime("%Y/%m/%d")


def abort(msg):
    log(msg, "ERROR")
    log("Script aborted. Fix the issues above and re-run.", "ERROR")
    sys.exit(1)


def close_workbook_safe(wb):
    if wb is None:
        return
    archive = getattr(wb, "_archive", None)
    try:
        wb.close()
    except Exception:
        pass

    if archive is not None:
        try:
            if hasattr(archive, "fp") and archive.fp is not None and getattr(archive.fp, "closed", False):
                archive.fp = None
        except Exception:
            pass
        try:
            archive.close()
        except Exception:
            pass

    try:
        if hasattr(wb, "_archive"):
            wb._archive = None
    except Exception:
        pass


def _repair_macros_via_excel(xlsm_path):
    """
    Re-open a saved .xlsm in Excel via COM and save it again.

    This repairs internal workbook content-type/relationship metadata that
    can occasionally be altered when saving through openpyxl, which may leave
    macros present but disconnected.
    """
    try:
        import win32com.client  # type: ignore[import-untyped]
        import pythoncom        # type: ignore[import-untyped]
    except ImportError:
        log(
            "pywin32 not installed — macro repair skipped. "
            "Install it with 'pip install pywin32' to preserve macros automatically.",
            "WARN",
        )
        return False

    abs_path = os.path.abspath(xlsm_path)
    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(abs_path)
        workbook.Save()
        log("Macros repaired — workbook re-saved via Excel COM", "OK")
        return True
    except Exception as exc:
        log(f"Excel COM macro repair failed: {exc}", "WARN")
        return False
    finally:
        if workbook:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def _write_sheets_via_com(output_path, transformed, group_id=None):
    """
    Write DataFrame data into the .xlsm using Excel COM automation.

    Excel performs the save, so workbook shapes and controls in the
    Instructions sheet remain intact.

    Returns True on success, False when pywin32 is not installed or COM write fails.
    """
    try:
        import win32com.client  # type: ignore[import-untyped]
        import pythoncom        # type: ignore[import-untyped]
    except ImportError:
        return False

    abs_path = os.path.abspath(output_path)
    _com_initialized = False
    excel = None
    workbook = None

    try:
        pythoncom.CoInitialize()
        _com_initialized = True

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = excel.Workbooks.Open(abs_path)

        for prop, value in (("ScreenUpdating", False), ("EnableEvents", False), ("Calculation", -4135)):
            try:
                setattr(excel, prop, value)
            except Exception as exc:
                log(f"Could not set Application.{prop} ({exc}); continuing.", "WARN")

        log("Writing data into FBDI template sheets (Excel COM) ...", "HEADER")

        for key, fbdi_df in transformed.items():
            sheet_name = SHEET_NAMES.get(key)
            if not sheet_name:
                continue

            ws = None
            for i in range(1, workbook.Sheets.Count + 1):
                if workbook.Sheets(i).Name == sheet_name:
                    ws = workbook.Sheets(i)
                    break

            if ws is None:
                context = format_context(group_id=group_id, sheet_name=sheet_name)
                log(
                    f"Sheet '{sheet_name}' not found in template. "
                    f"Check SHEET_NAMES config. Skipping.",
                    "WARN",
                    context,
                )
                continue

            used = ws.UsedRange
            scan_rows = min(20, used.Rows.Count)
            scan_cols = used.Columns.Count

            scan_data = ws.Range(ws.Cells(1, 1), ws.Cells(scan_rows, scan_cols)).Value
            if scan_data is None:
                scan_data = []
            elif not isinstance(scan_data[0], (tuple, list)):
                scan_data = (scan_data,)

            header_row_num = None
            for row_i, row_vals in enumerate(scan_data):
                if row_vals is None:
                    continue
                for cell_val in row_vals:
                    if cell_val and isinstance(cell_val, str):
                        if "business unit" in str(cell_val).strip().lower():
                            header_row_num = row_i + 1
                            break
                if header_row_num:
                    break

            if not header_row_num:
                header_row_num = 4
                context = format_context(group_id=group_id, sheet_name=sheet_name)
                log(
                    f"[{sheet_name}] Could not detect header row — defaulting to row 4.",
                    "WARN",
                    context,
                )

            header_row_vals = scan_data[header_row_num - 1] if scan_data else ()
            template_headers = {}
            for col_i, cell_val in enumerate(header_row_vals):
                if cell_val:
                    header_name = str(cell_val).strip()
                    if header_name and header_name not in template_headers:
                        template_headers[header_name] = col_i + 1

            data_start_row = header_row_num + 1
            last_used_row = used.Row + used.Rows.Count - 1
            if last_used_row >= data_start_row:
                ws.Range(
                    ws.Cells(data_start_row, 1),
                    ws.Cells(last_used_row, scan_cols),
                ).ClearContents()

            num_rows = len(fbdi_df)
            if num_rows == 0:
                log(f"[{sheet_name}] 0 rows written", "OK")
                continue

            max_col = max(template_headers.values()) if template_headers else 1

            mapped_cols = {col for col in fbdi_df.columns if col in template_headers}
            data_end_row = data_start_row + num_rows - 1
            mapped_col_nums = [
                col_num for col_name, col_num in template_headers.items()
                if col_name in mapped_cols
            ]
            if mapped_col_nums:
                try:
                    addrs = [
                        f"{_excel_col_letter(c)}{data_start_row}:{_excel_col_letter(c)}{data_end_row}"
                        for c in mapped_col_nums
                    ]
                    ws.Range(",".join(addrs)).NumberFormat = "@"
                except Exception:
                    for c in mapped_col_nums:
                        ws.Range(
                            ws.Cells(data_start_row, c),
                            ws.Cells(data_end_row, c),
                        ).NumberFormat = "@"

            bulk = [[None] * max_col for _ in range(num_rows)]
            for fbdi_col, series in fbdi_df.items():
                col_num = template_headers.get(fbdi_col)
                if col_num is None:
                    continue
                col_idx = col_num - 1
                clean_strs = series.fillna("").astype(str).tolist()
                for row_i, s in enumerate(clean_strs):
                    if s:
                        bulk[row_i][col_idx] = "'" + s

            ws.Range(
                ws.Cells(data_start_row, 1),
                ws.Cells(data_start_row + num_rows - 1, max_col),
            ).Value = bulk

            log(f"[{sheet_name}] {num_rows} rows written", "OK")

        workbook.Save()
        log(f"Workbook saved: {output_path}", "OK")
        return True

    except Exception as exc:
        import traceback
        log(f"Excel COM write error: {exc}", "ERROR")
        log(traceback.format_exc(), "ERROR")
        return False

    finally:
        if excel:
            for prop, value in (("Calculation", -4105), ("EnableEvents", True), ("ScreenUpdating", True)):
                try:
                    setattr(excel, prop, value)
                except Exception:
                    pass
        if workbook:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        if _com_initialized:
            pythoncom.CoUninitialize()


# =============================================================================
# CONFIGURATION LOADING
# =============================================================================

def load_source_file_config():
    """Load source_config.json and return config objects for mapping behavior."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, CONFIG_FILE)

    if not os.path.exists(config_path):
        abort(
            f"Source config file not found: {os.path.abspath(config_path)}\n"
            f"Please create '{CONFIG_FILE}' in the script folder."
        )

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception as exc:
        abort(f"Failed to read '{CONFIG_FILE}': {exc}")

    if not isinstance(config, dict):
        abort(f"'{CONFIG_FILE}' must contain a JSON object at the top level.")

    source_dir = config.get("source_dir", SOURCE_DIR)
    if not isinstance(source_dir, str) or not source_dir.strip():
        abort(f"'{CONFIG_FILE}' optional 'source_dir' must be a non-empty string.")
    source_dir = source_dir.strip()
    if not os.path.isabs(source_dir):
        source_dir = os.path.join(script_dir, source_dir)

    source_file_prefixes = config.get("source_file_prefixes", DEFAULT_SOURCE_FILE_PREFIXES)
    if not isinstance(source_file_prefixes, dict):
        abort(f"'{CONFIG_FILE}' 'source_file_prefixes' must be a JSON object.")

    missing = set(SOURCE_KEYS) - set(source_file_prefixes.keys())
    if missing:
        abort(f"'{CONFIG_FILE}' must define source_file_prefixes for: {sorted(missing)}.")

    column_map = config.get("column_map", {})
    if not isinstance(column_map, dict):
        abort(f"'{CONFIG_FILE}' optional 'column_map' must be a JSON object.")

    profiles = config.get("profiles", {})
    if not isinstance(profiles, dict):
        abort(f"'{CONFIG_FILE}' optional 'profiles' must be a JSON object.")

    for profile_name, mapping_by_key in profiles.items():
        if not isinstance(profile_name, str) or not profile_name.strip():
            abort(f"'{CONFIG_FILE}' profiles keys must be non-empty strings.")
        if not isinstance(mapping_by_key, dict):
            abort(f"'{CONFIG_FILE}' profiles['{profile_name}'] must be a JSON object.")
        for key, mapping in mapping_by_key.items():
            if not isinstance(mapping, dict):
                abort(
                    f"'{CONFIG_FILE}' profiles['{profile_name}']['{key}'] "
                    f"must be a JSON object."
                )

    source_column_map = config.get("source_column_map", {})
    if not isinstance(source_column_map, dict):
        abort(f"'{CONFIG_FILE}' optional 'source_column_map' must be a JSON object.")

    normalized_source_column_map = {}
    for source_name, value in source_column_map.items():
        if not isinstance(source_name, str) or not source_name.strip():
            abort(f"'{CONFIG_FILE}' source_column_map keys must be non-empty strings.")

        if isinstance(value, str):
            profile_name = value.strip()
            if profile_name not in profiles:
                abort(
                    f"'{CONFIG_FILE}' source_column_map['{source_name}'] references "
                    f"unknown profile '{profile_name}'. "
                    f"Defined profiles: {sorted(profiles)}"
                )
            mapping_by_key = profiles[profile_name]
        elif isinstance(value, dict):
            for key, mapping in value.items():
                if not isinstance(mapping, dict):
                    abort(
                        f"'{CONFIG_FILE}' source_column_map['{source_name}']['{key}'] "
                        f"must be a JSON object."
                    )
            mapping_by_key = value
        else:
            abort(
                f"'{CONFIG_FILE}' source_column_map['{source_name}'] must be either "
                f"a profile name (string) or a JSON object of per-key mappings."
            )

        normalized_source_column_map[source_name.strip().upper()] = mapping_by_key

    return source_dir, source_file_prefixes, column_map, normalized_source_column_map


def get_source_name_from_group_id(group_id):
    """Extract source system token from group id (e.g. GENEVA from GENEVA_INV_0001)."""
    if not group_id:
        return ""
    return clean_str(group_id).split("_", 1)[0].upper()


def apply_source_specific_overrides(column_map, source_name, source_overrides):
    """
    Apply source-specific overrides without affecting unrelated default mappings.

    For TRX/TRX_DIST mappings, only flexfield target-column families are
    source-controlled; all other mappings remain as defined in column_map.
    """
    if not source_overrides:
        return column_map

    for key, override_mapping in source_overrides.items():
        existing = dict(column_map.get(key, {}))
        override_mapping = dict(override_mapping)

        if key == "trx":
            existing = {
                src_col: tpl_col
                for src_col, tpl_col in existing.items()
                if normalize_column_name(tpl_col) not in CONTROLLED_TRX_TEMPLATE_COLUMNS
            }
        elif key == "trx_dist":
            existing = {
                src_col: tpl_col
                for src_col, tpl_col in existing.items()
                if normalize_column_name(tpl_col) not in CONTROLLED_TRX_DIST_TEMPLATE_COLUMNS
            }

        existing.update(override_mapping)
        column_map[key] = existing

    return column_map


# =============================================================================
# FILE DISCOVERY
# =============================================================================

def parse_source_filename(filename, prefix):
    """
    Parse a source filename and return the batch identifier if it matches the prefix.

    Pattern: {PREFIX}_{SOURCE}_{TYPE}_{SEQ}_{DATE}_{TIME}.csv
    Returns:  "{SOURCE}_{TYPE}_{SEQ}" string, or None if the file does not match.
    """
    if not filename.lower().endswith(".csv"):
        return None
    if not filename.startswith(prefix + "_"):
        return None

    stem = filename[:-4]
    remainder = stem[len(prefix) + 1:]
    parts = remainder.split("_")

    # Minimum: SOURCE, TYPE, SEQ, DATE, TIME
    if len(parts) < 5:
        return None

    # parts[1] must be a valid transaction type (INV / CM / DM)
    if parts[1].upper() not in VALID_TRANSACTION_TYPES:
        return None

    return f"{parts[0]}_{parts[1]}_{parts[2]}"


def discover_source_groups(source_dir, prefixes):
    """
    Scan source_dir and pair TRX / TRX_DIST files by batch identifier.
    Returns: {group_id: {"trx": full_path, "trx_dist": full_path}}
    """
    if not os.path.isdir(source_dir):
        abort(f"Source directory not found: {os.path.abspath(source_dir)}")

    groups = {}
    # Check longer prefixes first so TRX_DIST is matched before TRX
    sorted_prefixes = sorted(prefixes.items(), key=lambda item: -len(item[1]))

    for filename in sorted(os.listdir(source_dir)):
        for key, prefix in sorted_prefixes:
            batch_id = parse_source_filename(filename, prefix)
            if not batch_id:
                continue
            group_files = groups.setdefault(batch_id, {})
            if key in group_files:
                abort(
                    f"Duplicate file for batch '{batch_id}' key '{key}': "
                    f"'{filename}' vs '{os.path.basename(group_files[key])}'."
                )
            group_files[key] = os.path.join(source_dir, filename)
            break

    return groups


def format_source_groups(groups):
    lines = []
    for group_id, files in sorted(groups.items()):
        found = set(files.keys())
        status = "COMPLETE" if found == set(SOURCE_KEYS) else "INCOMPLETE"
        lines.append(
            f"  {group_id}: {status} ({len(files)}/{len(SOURCE_KEYS)} files) "
            f"[{', '.join(sorted(found))}]"
        )
    return "\n".join(lines)


# =============================================================================
# TEMPLATE HANDLING
# =============================================================================

def locate_template():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(script_dir, TEMPLATE_DIR, TEMPLATE_FILENAME)
    if not os.path.exists(template_path):
        abort(
            f"FBDI template not found at: {os.path.abspath(template_path)}\n"
            f"Please place '{TEMPLATE_FILENAME}' in the '{TEMPLATE_DIR}/' folder."
        )
    log(f"Template found: {template_path}", "OK")
    return template_path


def find_template_header_row(ws):
    """Locate the header row in an Oracle FBDI sheet (looks for 'Business Unit')."""
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if not row:
            continue
        normalized = [
            str(cell).strip().lower() if cell is not None else "" for cell in row
        ]
        if any("business unit" in cell for cell in normalized if cell):
            return [str(cell).strip() if cell is not None else "" for cell in row]

    # Fallback: first non-empty row
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        if any(cell is not None and str(cell).strip() for cell in row):
            return [str(cell).strip() if cell is not None else "" for cell in row]
    return []


def load_template_headers(template_path):
    """Return {key: [header_col_name, ...]} for each target sheet."""
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    headers = {}
    try:
        for key, sheet_name in SHEET_NAMES.items():
            if sheet_name not in wb.sheetnames:
                abort(
                    f"Template sheet '{sheet_name}' not found in workbook. "
                    f"Check SHEET_NAMES config."
                )
            ws = wb[sheet_name]
            header_row = find_template_header_row(ws)
            if not header_row:
                abort(f"Could not determine header row for template sheet '{sheet_name}'.")
            headers[key] = [cell for cell in header_row if cell]
            log(
                f"[{sheet_name}] Template header loaded: {len(headers[key])} columns", "OK"
            )
    finally:
        close_workbook_safe(wb)
    return headers


# =============================================================================
# COLUMN MAPPING
# =============================================================================

def build_column_map(frames, template_headers, source_paths=None, group_id=None):
    """
    Auto-build source CSV → template column mappings via name normalization.
    Columns whose normalized names do not match any template header are logged as
    unmapped warnings (e.g. trx_id which has no corresponding template field).
    """
    log("Building source-to-template column mappings ...", "HEADER")
    column_map = {}

    for key, sheet_name in SHEET_NAMES.items():
        df = frames.get(key)
        if df is None:
            log(f"[{key}] No source frame for '{sheet_name}'. Skipping.", "WARN")
            continue

        template_cols = template_headers.get(key, [])
        template_lookup = {
            normalize_column_name(col): col for col in template_cols if col
        }

        mapping = {}
        unmapped_src = []
        for src_col in df.columns:
            norm = normalize_column_name(src_col)
            if norm in template_lookup:
                mapping[src_col] = template_lookup[norm]
            else:
                unmapped_src.append(src_col)

        if unmapped_src:
            log(
                f"[{key}] {len(unmapped_src)} source column(s) not auto-mapped: {unmapped_src}",
                "WARN",
                format_context(group_id=group_id, key=key),
            )

        template_unmapped = [
            col for col in template_cols
            if col and normalize_column_name(col) not in
            {normalize_column_name(v) for v in mapping.values()}
        ]
        if template_unmapped:
            log(
                f"[{key}] {len(template_unmapped)} template column(s) not covered by source "
                f"(will be blank in output)",
                "WARN",
                format_context(group_id=group_id, key=key),
            )

        column_map[key] = mapping
        log(
            f"[{key}] auto-mapped {len(mapping)}/{len(df.columns)} source columns",
            "OK",
        )

    return column_map


def validate_column_map(column_map, frames, template_headers, source_paths=None, group_id=None):
    """Validate that all entries in the config column_map are resolvable."""
    log("Validating configured column map ...", "HEADER")
    error_count = 0

    for key, mapping in column_map.items():
        df = frames.get(key)
        source_path = source_paths.get(key) if source_paths else None
        context = format_context(group_id=group_id, file_path=source_path, key=key)

        if df is None:
            log(f"Column map key '{key}' exists in config but no source data is available.", "ERROR", context)
            error_count += 1
            continue

        template_cols = template_headers.get(key, [])
        source_lookup = {normalize_column_name(col): col for col in df.columns if col}
        template_lookup = {normalize_column_name(col): col for col in template_cols if col}

        for src_col, tpl_col in mapping.items():
            if normalize_column_name(src_col) not in source_lookup:
                log(
                    f"Config source column '{src_col}' not found in CSV for '{key}'.",
                    "ERROR", context,
                )
                error_count += 1
            if normalize_column_name(tpl_col) not in template_lookup:
                log(
                    f"Config template column '{tpl_col}' not found in sheet '{SHEET_NAMES.get(key)}'.",
                    "ERROR", context,
                )
                error_count += 1

    return error_count


def resolve_column_map_template_names(column_map, template_headers):
    """Resolve mapped template column names to the exact strings found in the workbook header."""
    resolved = {}
    for key, mapping in column_map.items():
        template_cols = template_headers.get(key, [])
        normalized_lookup = {
            normalize_column_name(col): col for col in template_cols if col
        }
        resolved[key] = {
            src_col: normalized_lookup.get(normalize_column_name(tpl_col), tpl_col)
            for src_col, tpl_col in mapping.items()
        }
    return resolved


# =============================================================================
# VALIDATION
# =============================================================================

def validate_trx_id_references(frames, source_paths=None, group_id=None):
    """
    Verify that every trx_id in TRX_DIST has a matching row in TRX.
    This ensures all distributions belong to a known invoice line.
    """
    log("Validating trx_id references (TRX_DIST -> TRX) ...", "HEADER")

    trx_df   = frames.get("trx")
    dist_df  = frames.get("trx_dist")

    if trx_df is None or dist_df is None:
        log("Cannot validate trx_id references — missing TRX or TRX_DIST data.", "WARN")
        return 0

    if TRX_ID_COLUMN not in trx_df.columns:
        log(f"Column '{TRX_ID_COLUMN}' not found in TRX file. Skipping trx_id validation.", "WARN")
        return 0

    if TRX_ID_COLUMN not in dist_df.columns:
        log(f"Column '{TRX_ID_COLUMN}' not found in TRX_DIST file. Skipping trx_id validation.", "WARN")
        return 0

    trx_ids      = set(trx_df[TRX_ID_COLUMN].apply(clean_str).tolist()) - {""}
    dist_trx_ids = dist_df[TRX_ID_COLUMN].apply(clean_str)

    unmatched = dist_df[
        (dist_trx_ids != "") & (~dist_trx_ids.isin(trx_ids))
    ]

    if not unmatched.empty:
        missing_ids = unmatched[TRX_ID_COLUMN].unique().tolist()
        context = format_context(
            group_id=group_id,
            file_path=source_paths.get("trx_dist") if source_paths else None,
            key="trx_dist",
        )
        log(
            f"TRX_DIST has {len(unmatched)} row(s) with trx_id not present in TRX: {missing_ids[:10]}",
            "ERROR", context,
        )
        return len(unmatched)

    log(f"All {len(dist_df)} TRX_DIST trx_id values matched in TRX.", "OK")
    return 0


def validate_mandatory_star_fields(transformed, template_headers, group_id=None):
    """Check that every mandatory (*) template column has no blank values."""
    log("Validating mandatory '*' template fields ...", "HEADER")
    error_count = 0

    for key, df in transformed.items():
        template_cols = template_headers.get(key, [])
        mandatory_cols = [col for col in template_cols if col.startswith("*")]
        if not mandatory_cols:
            continue

        context = format_context(group_id=group_id, key=key)
        output_lookup = {normalize_column_name(col): col for col in df.columns}

        for tpl_col in mandatory_cols:
            output_col = output_lookup.get(normalize_column_name(tpl_col))
            if output_col is None:
                log(
                    f"Mandatory template column '{tpl_col}' is missing from transformed output.",
                    "ERROR", context,
                )
                error_count += 1
                continue

            def is_blank(v):
                if v is None:
                    return True
                if isinstance(v, float) and pd.isna(v):
                    return True
                return str(v).strip() == ""

            blank_mask = df[output_col].apply(is_blank)
            if blank_mask.any():
                excel_rows = [i + 2 for i in df.index[blank_mask].tolist()]
                log(
                    f"Mandatory column '{tpl_col}' has blank values in "
                    f"{len(excel_rows)} row(s): Excel rows {excel_rows[:20]}",
                    "ERROR", context,
                )
                error_count += len(excel_rows)
            else:
                log(f"Mandatory column '{tpl_col}' — all populated", "OK", context)

    return error_count


def validate_date_columns(transformed, group_id=None):
    """Validate that every populated date column value is in YYYY/MM/DD format."""
    log("Validating date column formats ...", "HEADER")
    error_count = 0

    for key, df in transformed.items():
        date_cols = [col for col in df.columns if is_date_column(col)]
        if not date_cols:
            continue

        context = format_context(group_id=group_id, key=key)
        for col in date_cols:
            populated = df[col].apply(clean_str) != ""
            bad_mask = populated & df[col].apply(
                lambda v: not _DATE_FORMAT_RE.match(clean_str(v)) if clean_str(v) else False
            )
            if bad_mask.any():
                excel_rows = [i + 2 for i in df.index[bad_mask].tolist()]
                bad_vals = df.loc[bad_mask, col].tolist()
                log(
                    f"Column '{col}': {len(excel_rows)} value(s) not in YYYY/MM/DD "
                    f"format at Excel rows {excel_rows[:20]}: {bad_vals[:5]}",
                    "ERROR", context,
                )
                error_count += len(excel_rows)
            else:
                populated_count = int(populated.sum())
                if populated_count > 0:
                    log(
                        f"Column '{col}' — {populated_count} value(s) all in YYYY/MM/DD format",
                        "OK", context,
                    )

    return error_count


# =============================================================================
# READ SOURCES
# =============================================================================

def read_sources(source_files):
    """Read TRX and TRX_DIST CSV files into DataFrames."""
    log("Reading source CSV files ...", "HEADER")
    frames = {}
    for key, path in source_files.items():
        if not os.path.exists(path):
            abort(f"Source file not found: {os.path.abspath(path)}")
        df = pd.read_csv(path, dtype=str)
        df.columns = [c.strip() for c in df.columns]
        df = df.apply(lambda col: col.str.strip() if col.dtype == object else col)
        log(f"Loaded '{key}': {len(df)} rows | {len(df.columns)} columns", "OK")
        frames[key] = df
    return frames


# =============================================================================
# TRANSFORM
# =============================================================================

def transform(frames, column_map, template_headers, source_paths=None, group_id=None):
    """
    Apply column mapping to each source DataFrame.
    - '.' placeholder values are converted to empty/None.
    - Date columns receive Python datetime objects so Excel formats them correctly.
    - Source columns absent from the mapping (e.g. trx_id) are dropped.
    """
    log("Transforming data to FBDI column structure ...", "HEADER")
    transformed = {}

    for key, sheet_name in SHEET_NAMES.items():
        source_path = source_paths.get(key) if source_paths else None
        context = format_context(group_id=group_id, file_path=source_path, key=key)

        df = frames.get(key)
        if df is None:
            log(f"[{key}] No source data for '{sheet_name}'. Skipping.", "WARN", context)
            continue

        # Exclude REC (receivable control) rows from distributions — Oracle
        # populates these automatically and they must not be loaded via FBDI.
        if key == "trx_dist" and "account_class" in df.columns:
            rec_mask = df["account_class"].str.strip().str.upper() == "REC"
            excluded = rec_mask.sum()
            if excluded:
                log(f"[{key}] Excluded {excluded} REC account_class row(s).", "OK", context)
                df = df[~rec_mask].reset_index(drop=True)

        mapping = column_map.get(key, {})
        out = pd.DataFrame(index=df.index)

        for src_col, fbdi_col in mapping.items():
            if src_col not in df.columns:
                log(
                    f"[{key}] Source column '{src_col}' not in CSV — "
                    f"'{fbdi_col}' will be blank.",
                    "WARN", context,
                )
                out[fbdi_col] = None
                continue

            if is_date_column(fbdi_col):
                out[fbdi_col] = df[src_col].apply(parse_to_date_value)
            else:
                out[fbdi_col] = df[src_col].apply(clean_str).replace("", None)

        transformed[key] = out
        log(f"[{key}] Transformed: {len(out)} rows, {len(out.columns)} columns", "OK")

    return transformed


# =============================================================================
# WRITE TO TEMPLATE
# =============================================================================

def write_to_template(template_path, transformed, batch_id, group_id=None):
    """
    Copy the Oracle FBDI .xlsm template to the output folder, then write
    each transformed DataFrame into the corresponding sheet starting below
    the header row. Returns the path to the written output file.
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(script_dir, OUTPUT_DIR)
    os.makedirs(out_dir, exist_ok=True)

    base_name = os.path.splitext(TEMPLATE_FILENAME)[0]
    output_name = f"{base_name}_{batch_id}.xlsm"
    output_path = os.path.join(out_dir, output_name)

    shutil.copy2(template_path, output_path)
    log(f"Template copied to: {output_path}", "OK")

    if _write_sheets_via_com(output_path, transformed, group_id):
        return output_path

    wb = openpyxl.load_workbook(output_path, keep_vba=True)
    try:
        log("Writing data into FBDI template sheets ...", "HEADER")

        for key, fbdi_df in transformed.items():
            sheet_name = SHEET_NAMES.get(key)
            if not sheet_name:
                continue

            if sheet_name not in wb.sheetnames:
                log(
                    f"Sheet '{sheet_name}' not found in template. Skipping.",
                    "WARN",
                    format_context(group_id=group_id, sheet_name=sheet_name),
                )
                continue

            ws = wb[sheet_name]

            # Locate the header row by looking for "Business Unit" which only
            # appears in the actual column header row, not in title/instruction rows.
            header_row_num = None
            for row in ws.iter_rows(min_row=1, max_row=20):
                for cell in row:
                    if cell.value and "business unit" in str(cell.value).strip().lower():
                        header_row_num = cell.row
                        break
                if header_row_num:
                    break

            if not header_row_num:
                # Oracle Invoice FBDI templates always put headers at row 4
                header_row_num = 4
                log(
                    f"[{sheet_name}] Could not detect header row — defaulting to row 4.",
                    "WARN",
                )

            # Build header-name → column-number index (keep first occurrence only)
            template_col_index = {}
            for cell in ws[header_row_num]:
                if cell.value:
                    name = str(cell.value).strip()
                    if name and name not in template_col_index:
                        template_col_index[name] = cell.column

            # Clear any pre-existing data rows below the header
            data_start_row = header_row_num + 1
            for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

            # Write transformed rows
            rows_written = 0
            for _, df_row in fbdi_df.iterrows():
                excel_row = data_start_row + rows_written
                for fbdi_col, value in df_row.items():
                    col_num = template_col_index.get(fbdi_col)
                    if col_num is None:
                        continue
                    if value is None or pd.isna(value) or value == "":
                        ws.cell(row=excel_row, column=col_num, value=None)
                    else:
                        cell = ws.cell(row=excel_row, column=col_num, value=str(value))
                        cell.number_format = "@"
                        cell.quote_prefix = True
                rows_written += 1

            log(f"[{sheet_name}] {rows_written} rows written", "OK")

        wb.save(output_path)
        log(f"Workbook saved: {output_path}", "OK")
        _repair_macros_via_excel(output_path)
        return output_path

    finally:
        close_workbook_safe(wb)


# =============================================================================
# PROCESS A SINGLE BATCH GROUP
# =============================================================================

def process_source_group(
    group_id,
    source_files,
    template_path,
    config_column_map,
    config_source_column_map,
):
    """
    End-to-end processing for one batch group:
      1. Read CSV sources
      2. Auto-build column map, then merge any explicit overrides from config
      3. Validate column map and trx_id references
      4. Transform data
      5. Validate mandatory fields
      6. Write output file
      7. Print summary
    """
    log(f"Processing batch group: {group_id}", "HEADER")

    frames = read_sources(source_files)
    template_headers = load_template_headers(template_path)

    column_map = build_column_map(frames, template_headers, source_paths=source_files, group_id=group_id)

    # Merge global config overrides on top of auto-detected mapping
    if config_column_map:
        for key, mapping in config_column_map.items():
            if key in column_map and isinstance(column_map[key], dict):
                column_map[key].update(mapping)
            else:
                column_map[key] = mapping
        log("Config column_map overrides merged.", "OK")

    # Apply source-specific overrides while preserving unrelated defaults.
    source_name = get_source_name_from_group_id(group_id)
    source_overrides = config_source_column_map.get(source_name, {})
    if source_overrides:
        column_map = apply_source_specific_overrides(column_map, source_name, source_overrides)
        log(
            f"Source-specific overrides applied for source '{source_name}' "
            f"on keys: {sorted(source_overrides.keys())}",
            "OK",
        )

    column_map = resolve_column_map_template_names(column_map, template_headers)

    error_count  = validate_column_map(
        column_map, frames, template_headers, source_paths=source_files, group_id=group_id
    )
    error_count += validate_trx_id_references(frames, source_paths=source_files, group_id=group_id)

    transformed  = transform(frames, column_map, template_headers, source_paths=source_files, group_id=group_id)
    error_count += validate_mandatory_star_fields(transformed, template_headers, group_id=group_id)
    error_count += validate_date_columns(transformed, group_id=group_id)

    output_path  = write_to_template(template_path, transformed, group_id)
    print_summary(group_id, transformed, output_path, error_count)
    return output_path, error_count


# =============================================================================
# SUMMARY
# =============================================================================

def print_summary(batch_id, transformed, output_path, error_count):
    log("", "HEADER")
    log("=" * 62, "HEADER")
    log(f" Invoice FBDI Preparation Summary — Batch: {batch_id}", "HEADER")
    log("=" * 62, "HEADER")

    total = 0
    for key, df in transformed.items():
        sheet = SHEET_NAMES.get(key, key)
        log(f"  {sheet:<40} {len(df):>5} rows", "HEADER")
        total += len(df)

    log(f"  {'TOTAL':<40} {total:>5} rows", "HEADER")
    log("=" * 62, "HEADER")

    if error_count > 0:
        log(
            f"  {error_count} validation issue(s) found — review warnings above.",
            "HEADER"
        )
        log(
            "  Output file was still produced. Review before uploading to Fusion.",
            "HEADER"
        )
    else:
        log("  All validations passed — no issues found.", "HEADER")

    log("", "HEADER")
    log(f"  Output file: {os.path.abspath(output_path)}", "HEADER")
    log("", "HEADER")
    log("  Next steps:", "HEADER")
    log("  1. Open the output .xlsm and spot-check a sample of rows", "HEADER")
    log("  2. In Fusion: Tools -> File Import and Export -> Upload the .xlsm", "HEADER")
    log("  3. Run scheduled process: Import AutoInvoice", "HEADER")
    log("  4. Review the ESS log and error report in Fusion", "HEADER")
    log("=" * 62, "HEADER")


# =============================================================================
# ARGUMENT PARSING
# =============================================================================

def parse_arguments():
    parser = argparse.ArgumentParser(
        description=(
            "Prepare Oracle Fusion AR Invoice FBDI from TRX / TRX_DIST source files. "
            "Every complete TRX + TRX_DIST pair in source_data/ is processed automatically."
        )
    )
    parser.add_argument(
        "--group",
        dest="group_id",
        help="Process only this batch identifier, e.g. DEFAKTO_INV_0001.",
    )
    parser.add_argument(
        "--list-groups",
        action="store_true",
        help="List discovered source groups and exit.",
    )
    parser.add_argument(
        "--parallel",
        action="store_true",
        help="Process all complete source groups in parallel.",
    )
    parser.add_argument(
        "--max-workers",
        type=int,
        default=min(4, os.cpu_count() or 1),
        help="Maximum parallel workers when using --parallel.",
    )
    return parser.parse_args()


# =============================================================================
# MAIN
# =============================================================================

def main():
    args = parse_arguments()

    log("=" * 62, "HEADER")
    log(" Oracle Fusion AR — Invoice FBDI Preparation", "HEADER")
    log("=" * 62, "HEADER")

    template_path = locate_template()
    (
        source_dir,
        source_prefixes,
        config_column_map,
        config_source_column_map,
    ) = load_source_file_config()

    groups = discover_source_groups(source_dir, source_prefixes)
    if not groups:
        abort(f"No source files found in '{os.path.abspath(source_dir)}'.")

    complete_groups = {
        gid: files
        for gid, files in groups.items()
        if set(files.keys()) == set(SOURCE_KEYS)
    }

    incomplete = {gid for gid in groups if gid not in complete_groups}
    if incomplete:
        log(
            f"Skipping {len(incomplete)} incomplete group(s) (missing TRX or TRX_DIST): "
            f"{sorted(incomplete)}",
            "WARN",
        )

    if args.list_groups:
        log("Discovered source groups:", "HEADER")
        print(format_source_groups(groups))
        return

    if not complete_groups:
        abort(
            "No complete source groups to process "
            "(each batch needs both a TRX and TRX_DIST file).\n"
            f"{format_source_groups(groups)}"
        )

    log(f"Found {len(complete_groups)} complete group(s) to process.", "OK")

    if args.parallel:
        if args.group_id:
            abort("--group cannot be used together with --parallel.")
        log(
            f"Processing {len(complete_groups)} group(s) in parallel "
            f"with {args.max_workers} workers.",
            "HEADER",
        )
        with concurrent.futures.ProcessPoolExecutor(max_workers=args.max_workers) as executor:
            futures = {
                executor.submit(
                    process_source_group,
                    gid, files, template_path, config_column_map, config_source_column_map,
                ): gid
                for gid, files in complete_groups.items()
            }
            total_errors = 0
            for future in concurrent.futures.as_completed(futures):
                gid = futures[future]
                try:
                    _, errors = future.result()
                    total_errors += errors
                except Exception as exc:
                    abort(f"Processing group '{gid}' failed: {exc}")
        if total_errors:
            abort(f"Processing completed with {total_errors} validation issue(s).")
        return

    if args.group_id:
        if args.group_id not in complete_groups:
            abort(
                f"Requested group '{args.group_id}' not found or is incomplete.\n"
                f"Available complete groups:\n{format_source_groups(complete_groups)}"
            )
        process_source_group(
            args.group_id,
            complete_groups[args.group_id],
            template_path,
            config_column_map,
            config_source_column_map,
        )
        return

    # Process all complete groups sequentially
    total_errors = 0
    for gid, files in sorted(complete_groups.items()):
        _, errors = process_source_group(
            gid,
            files,
            template_path,
            config_column_map,
            config_source_column_map,
        )
        total_errors += errors

    if total_errors:
        abort(f"Processing completed with {total_errors} validation issue(s). See logs above.")


if __name__ == "__main__":
    main()
